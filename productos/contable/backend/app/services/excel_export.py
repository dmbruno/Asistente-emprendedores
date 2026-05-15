"""Exportación a Excel con openpyxl: facturas del periodo del cliente."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def generar_excel_periodo(facturas: list[dict[str, Any]], mes: int, anio: int) -> bytes:
    """Devuelve un Excel en bytes con las facturas del periodo."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{mes:02d}-{anio}"

    headers = [
        "Fecha", "Tipo", "Comprobante", "Punto Venta", "Número",
        "CUIT Contraparte", "Razón Social", "Subtotal", "IVA", "Total", "Estado",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for f in facturas:
        ws.append(
            [
                f.get("fecha"),
                f.get("tipo"),
                f.get("tipo_comprobante"),
                f.get("punto_venta"),
                f.get("numero"),
                f.get("cuit_contraparte"),
                f.get("razon_social_contraparte"),
                float(f.get("subtotal") or 0),
                float(f.get("iva") or 0),
                float(f.get("total") or 0),
                f.get("estado"),
            ]
        )

    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
